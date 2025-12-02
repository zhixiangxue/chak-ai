"""
Excel spreadsheet attachment with default reader.
"""
from typing import Optional, Callable, Union, Awaitable
import os
import tempfile
from .base import Attachment, MimeType, detect_content_type, AttachmentContent, ContentType


def default_xlsx_reader(content: str) -> AttachmentContent:
    """
    Default synchronous XLSX reader.
    
    Args:
        content: Local file path or URL to XLSX
        
    Returns:
        AttachmentContent with extracted text and metadata
    """
    try:
        import openpyxl
        import urllib.request
        
        content_type = detect_content_type(content)
        
        # Handle different content types
        if content_type == ContentType.URL:
            # Download URL to temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                urllib.request.urlretrieve(content, tmp.name)
                file_path = tmp.name
        
        elif content_type == ContentType.LOCAL_PATH:
            # Local path: check if file exists
            if not os.path.exists(content):
                return AttachmentContent(content="", meta={"error": f"File not found: {content}"})
            file_path = content
        
        elif content_type == ContentType.TEXT:
            # Plain text: not a valid file path
            return AttachmentContent(content="", meta={"error": "Invalid content: expected file path or URL, got plain text"})
        
        else:  # ContentType.DATA_URI
            # Data URI: not supported for Excel
            return AttachmentContent(content="", meta={"error": "Data URI not supported for Excel files"})
        
        # Now file_path is guaranteed to be a valid local file
        wb = openpyxl.load_workbook(file_path, read_only=True)
        text_parts = []
        
        for sheet in wb.worksheets:
            text_parts.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)
        
        text = "\n".join(text_parts)
        meta = {
            "sheets": len(wb.worksheets),
            "sheet_names": [sheet.title for sheet in wb.worksheets],
        }
        
        # Clean up temp file if downloaded from URL
        if content_type == ContentType.URL:
            try:
                os.unlink(file_path)
            except Exception:
                pass
        
        return AttachmentContent(content=text, meta=meta)
    
    except Exception as e:
        return AttachmentContent(content="", meta={"error": str(e)})


async def default_xlsx_reader_async(content: str) -> AttachmentContent:
    """
    Default asynchronous XLSX reader.
    
    Args:
        content: Local file path or URL to XLSX
        
    Returns:
        AttachmentContent with extracted text and metadata
    """
    try:
        import openpyxl
        import aiohttp
        import aiofiles
        
        content_type = detect_content_type(content)
        
        # Handle different content types
        if content_type == ContentType.URL:
            # Download URL to temp file asynchronously
            async with aiohttp.ClientSession() as session:
                async with session.get(content) as response:
                    response.raise_for_status()
                    data = await response.read()
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(data)
                        file_path = tmp.name
        
        elif content_type == ContentType.LOCAL_PATH:
            # Local path: check if file exists
            if not os.path.exists(content):
                return AttachmentContent(content="", meta={"error": f"File not found: {content}"})
            file_path = content
        
        elif content_type == ContentType.TEXT:
            # Plain text: not a valid file path
            return AttachmentContent(content="", meta={"error": "Invalid content: expected file path or URL, got plain text"})
        
        else:  # ContentType.DATA_URI
            # Data URI: not supported for Excel
            return AttachmentContent(content="", meta={"error": "Data URI not supported for Excel files"})
        
        # Now file_path is guaranteed to be a valid local file
        wb = openpyxl.load_workbook(file_path, read_only=True)
        text_parts = []
        
        for sheet in wb.worksheets:
            text_parts.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)
        
        text = "\n".join(text_parts)
        meta = {
            "sheets": len(wb.worksheets),
            "sheet_names": [sheet.title for sheet in wb.worksheets],
        }
        
        # Clean up temp file if downloaded from URL
        if content_type == ContentType.URL:
            try:
                os.unlink(file_path)
            except Exception:
                pass
        
        return AttachmentContent(content=text, meta=meta)
    
    except Exception as e:
        return AttachmentContent(content="", meta={"error": str(e)})


class Excel(Attachment):
    """Microsoft Excel spreadsheet attachment (XLS/XLSX)"""
    
    def __init__(
        self, 
        source: str,
        reader: Optional[Callable[[str], Union[AttachmentContent, Awaitable[AttachmentContent]]]] = None
    ):
        """
        Create an Excel attachment.
        
        Args:
            source: Local file path or URL to XLS/XLSX file
            reader: Optional custom reader function (defaults to default_xlsx_reader)
        
        Examples:
            >>> Excel("data.xlsx")
            >>> Excel("https://example.com/spreadsheet.xlsx")
        """
        # Auto-detect XLSX vs XLS from extension
        mime_type = MimeType.XLSX if source.lower().endswith('.xlsx') else MimeType.XLS
        reader = reader or default_xlsx_reader_async
        super().__init__(source, mime_type, reader)
