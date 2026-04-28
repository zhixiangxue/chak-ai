"""
Excel: Built-in spreadsheet tool for chak

Provides three public methods exposed as LLM tools via NativeObjectTool:

    - sheets: List sheet names in an Excel workbook
    - read:   Read a sheet (or CSV) as a formatted table
    - write:  Write data to a sheet, replacing its contents

Supported formats:
    .xlsx  — via openpyxl (pip install openpyxl)
    .csv   — built-in csv module (zero extra deps)

Usage:
    from chak.tools.std import Excel
    excel = Excel()
    conv = Conversation(model, tools=[excel])
"""

import csv
from pathlib import Path
from typing import List, Optional

_MAX_ROWS = 500
_MAX_CELL_LEN = 300


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed — run `pip install openpyxl`")


def _fmt_table(headers: list, rows: list, max_rows: int) -> str:
    """Format spreadsheet data as a plain-text ASCII table."""
    if not rows:
        return f"(0 rows)\nColumns: {', '.join(str(h) for h in headers)}"

    total = len(rows)
    display = rows[:max_rows]
    n_cols = len(headers)

    # Compute column widths from headers + display rows
    widths = [min(len(str(h)), _MAX_CELL_LEN) for h in headers]
    for row in display:
        for i in range(n_cols):
            val = row[i] if i < len(row) else ""
            cell = "" if val is None else str(val)[:_MAX_CELL_LEN]
            widths[i] = max(widths[i], len(cell))

    def fmt_row(vals: list) -> str:
        parts = []
        for i in range(n_cols):
            val = vals[i] if i < len(vals) else ""
            cell = "" if val is None else str(val)[:_MAX_CELL_LEN]
            parts.append(cell.ljust(widths[i]))
        return " | ".join(parts)

    lines = [fmt_row(headers), "-+-".join("-" * w for w in widths)]
    lines += [fmt_row(r) for r in display]
    suffix = (
        f"\n\n(showing {max_rows} of {total}+ rows — pass max_rows to see more)"
        if total > max_rows
        else f"\n\n({total} row{'s' if total != 1 else ''})"
    )
    return "\n".join(lines) + suffix


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

class Excel:
    """Spreadsheet tool for reading and writing .xlsx and .csv files.

    All path arguments may be absolute or relative to workdir (if set).
    .xlsx requires openpyxl: ``pip install openpyxl``.
    .csv uses Python's built-in csv module — no extra packages needed.

    Example::

        excel = Excel()
        conv = Conversation(model, tools=[excel])
        await conv.asend("Read archived.xlsx and find rows not updated in 10+ days")
    """

    def __init__(
        self,
        workdir: Optional[str] = None,
        max_rows: int = _MAX_ROWS,
    ):
        """
        Args:
            workdir: Optional root directory. Relative paths resolve against it.
                     When set, all paths are restricted to this directory tree.
            max_rows: Maximum rows returned by read() (default 500).
        """
        self._workdir: Optional[Path] = Path(workdir).resolve() if workdir else None
        self._max_rows = max_rows

    def _resolve(self, path: str) -> Path:
        p = Path(path).expanduser()
        if not p.is_absolute() and self._workdir:
            p = self._workdir / p
        resolved = p.resolve()
        if self._workdir:
            try:
                resolved.relative_to(self._workdir)
            except ValueError:
                raise PermissionError(
                    f"Path '{path}' is outside workdir '{self._workdir}'"
                )
        return resolved

    # ------------------------------------------------------------------
    # sheets
    # ------------------------------------------------------------------

    def sheets(self, path: str) -> str:
        """List the sheet names in an Excel workbook.

        Call this first to discover available sheets before calling read().
        Not applicable to .csv files (a CSV is always a single sheet).

        Args:
            path: Path to .xlsx file (absolute or relative to workdir).

        Returns:
            Newline-separated list of sheet names, or an error string.
        """
        try:
            fp = self._resolve(path)
            if not fp.exists():
                return f"Error: File not found: {path}"
            if fp.suffix.lower() == ".csv":
                return "(CSV file — single implicit sheet, no named sheets)"
            openpyxl = _require_openpyxl()
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return "\n".join(names) if names else "(no sheets found)"
        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error listing sheets: {e}"

    # ------------------------------------------------------------------
    # read
    # ------------------------------------------------------------------

    def read(
        self,
        path: str,
        sheet: Optional[str] = None,
        max_rows: Optional[int] = None,
        skip_empty_rows: bool = True,
    ) -> str:
        """Read a spreadsheet sheet or CSV file and return it as a formatted table.

        The first non-empty row is treated as the header row.
        Use sheets() first to discover sheet names in .xlsx files.

        Args:
            path:            Path to .xlsx or .csv file.
            sheet:           Sheet name for .xlsx (default: first/active sheet).
            max_rows:        Maximum data rows to return (default from constructor).
            skip_empty_rows: Skip rows where all cells are empty (default True).

        Returns:
            Formatted table with header row, or an error string.
        """
        try:
            fp = self._resolve(path)
            if not fp.exists():
                return f"Error: File not found: {path}"
            n = max_rows if max_rows is not None else self._max_rows
            if fp.suffix.lower() == ".csv":
                return self._read_csv(fp, n, skip_empty_rows)
            return self._read_xlsx(fp, sheet, n, skip_empty_rows)
        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error reading: {e}"

    def _read_csv(self, fp: Path, max_rows: int, skip_empty: bool) -> str:
        with fp.open(encoding="utf-8-sig", errors="replace") as f:
            all_rows = [
                r for r in csv.reader(f)
                if not skip_empty or any(c.strip() for c in r)
            ]
        if not all_rows:
            return "(empty CSV file)"
        return _fmt_table(all_rows[0], all_rows[1:], max_rows)

    def _read_xlsx(
        self, fp: Path, sheet: Optional[str], max_rows: int, skip_empty: bool
    ) -> str:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            if ws is None:
                return (
                    f"Error: sheet '{sheet}' not found. "
                    f"Available: {', '.join(wb.sheetnames)}"
                )
            rows: list = []
            for row in ws.iter_rows(values_only=True):
                if skip_empty and all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append(list(row))
        finally:
            wb.close()
        if not rows:
            return "(empty sheet)"
        headers = [str(c) if c is not None else "" for c in rows[0]]
        return _fmt_table(headers, rows[1:], max_rows)

    # ------------------------------------------------------------------
    # write
    # ------------------------------------------------------------------

    def write(
        self,
        path: str,
        data: List[List],
        sheet: Optional[str] = None,
    ) -> str:
        """Write data to a sheet, replacing its contents (or creating the file).

        The first inner list is used as the header row.
        For .csv files the sheet argument is ignored.

        Args:
            path:  Path to .xlsx or .csv file. Created if it does not exist.
            data:  List of rows; each row is a list of cell values.
                   Example: [["name", "score"], ["Alice", 95], ["Bob", 87]]
            sheet: Sheet name for .xlsx (default: "Sheet1").

        Returns:
            Success message or error string.
        """
        try:
            fp = self._resolve(path)
            fp.parent.mkdir(parents=True, exist_ok=True)
            ext = fp.suffix.lower()

            if ext == ".csv":
                with fp.open("w", newline="", encoding="utf-8-sig") as f:
                    csv.writer(f).writerows(data)
                return f"Written {len(data)} rows to {fp}"

            openpyxl = _require_openpyxl()
            wb = openpyxl.load_workbook(str(fp)) if fp.exists() else openpyxl.Workbook()
            sheet_name = sheet or "Sheet1"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = wb.create_sheet(sheet_name)
                # Remove the default empty sheet if it exists
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                    del wb["Sheet"]
            for row in data:
                ws.append(row)
            wb.save(str(fp))
            return f"Written {len(data)} rows to sheet '{sheet_name}' in {fp}"
        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error writing: {e}"

    def __repr__(self) -> str:
        wd = str(self._workdir) if self._workdir else "unrestricted"
        return f"<Excel workdir={wd} max_rows={self._max_rows}>"
